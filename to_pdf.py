import os
import traceback
import pandas as pd

# hwp
import win32com.client as win32
import win32gui
# excel
import openpyxl as op
# pdf
import shutil
# doc
import comtypes.client
# text
import aspose.words as aw

# image
from PIL import Image


class ConvertPdf:
    from_path = None # 읽을 폴더 경로
    to_path = None # 쓸 폴더 경로
    all_files = None # 모든 파일들의 이름을 담은 리스트
    def __init__(self,from_path, to_path, d):
        self.from_path = from_path + "/" + d
        self.to_path = to_path + "/" + d

        self.all_files = os.listdir(self.from_path)
        self.error_dict = {"error_file":[],"error_message":[]}

        if not os.path.exists(self.to_path):
            os.mkdir(self.to_path)

        for file in self.all_files:
            print(file)
            file_type = os.path.splitext(file)[1][1:]
            file_type = file_type.lower()

            if file_type in ("pdf"):
                self.pdf2pdf(file)
            elif file_type in ('hwp'):
                self.hwp2pdf(file)
            elif file_type in ('png','jpg','jpeg','jfif',"bmp"):
                self.img2pdf(file)
            elif file_type in\
                    ('xlsx','xls'):
                self.exl2pdf_v2(file)
            elif file_type in ('txt'):
                self.text2pdf(file)
            elif file_type in ('docx'):
                self.word2pdf(file)
            elif file_type in ('ppt','pptx'):
                self.ppt2pdf(file)
            else:
                self.error_dict["error_file"].append(file)
                self.error_dict["error_message"].append("no type converter")

    def change_file_name_pdf(self,file_name,to_type="pdf"):
        L = file_name.split(".")
        L[-1]=to_type
        return ".".join(L)

    def to_csv_error_file(self,path):
        pd.DataFrame(self.error_dict).to_csv(path,encoding="utf-8",index=False)

    def text2pdf(self,file_name):
        try:
            doc = aw.Document(os.path.join(self.from_path, file_name))
            pdf_file_name = self.change_file_name_pdf(file_name)
            doc.save(os.path.join(self.to_path, pdf_file_name))
        except:
            self.error_dict["error_file"].append(file_name)
            self.error_dict["error_message"].append(traceback.format_exc())

    def word2pdf(self,file_name):
        try:
            word = comtypes.client.CreateObject('Word.Application')
            word.Visible = True

            doc = word.Documents.Open(os.path.join(self.from_path, file_name))
            pdf_file_name = self.change_file_name_pdf(file_name)
            doc.SaveAs(os.path.join(self.to_path, pdf_file_name), FileFormat=17)
            doc.Close()
            word.Quit()
        except:
            self.error_dict["error_file"].append(file_name)
            self.error_dict["error_message"].append(traceback.format_exc())

    def ppt2pdf(self,file_name):
        powerpoint = comtypes.client.CreateObject("Powerpoint.Application")
        powerpoint.Visible = True

        slides = powerpoint.Presentations.Open(os.path.join(self.from_path, file_name))
        pdf_file_name = self.change_file_name_pdf(file_name)
        full_to_path = os.path.join(self.to_path, pdf_file_name).replace("/", "\\")

        slides.SaveAs(full_to_path, FileFormat=32)
        slides.Close()

    def pdf2pdf(self,file_name):
        try:
            shutil.copyfile(os.path.join(self.from_path, file_name),os.path.join(self.to_path, file_name))
            # print(file_name, "=>", file_name)
        except:
            self.error_dict["error_file"].append(file_name)
            self.error_dict["error_message"].append(traceback.format_exc())

    def hwp2pdf(self, file_name):
        try:
            if file_name.endswith("hwpx"):
                hwp_file_name = self.change_file_name_pdf(file_name, "hwp")
                os.rename(os.path.join(self.from_path,file_name),os.path.join(self.from_path,hwp_file_name))
                file_name = hwp_file_name

            # hwp용
            hwp = win32.gencache.EnsureDispatch("HWPFrame.HwpObject")  # hwp 창열기
            hwp.RegisterModule('FilePathCheckDLL', 'AutomationModule')  # 보안모듈 삭제
            win32gui.FindWindow(None, 'Noname 1 - HWP')

            hwp.Open(os.path.join(self.from_path, file_name))
            pdf_file_name = self.change_file_name_pdf(file_name)
            hwp.SaveAs(os.path.join(self.to_path, pdf_file_name), "PDF")
            hwp.Quit()
            # print(file_name, "=>", pdf_file_name)
        except:
            self.error_dict["error_file"].append(file_name)
            self.error_dict["error_message"].append(traceback.format_exc())

    def img2pdf(self,file_name):
        try:
            im = Image.open(os.path.join(self.from_path,file_name)).convert("RGB")
            pdf_file_name = self.change_file_name_pdf(file_name)
            im.save(os.path.join(self.to_path,pdf_file_name),save_all=True)
            # print(file_name, "=>", pdf_file_name)
        except:
            self.error_dict["error_file"].append(file_name)
            self.error_dict["error_message"].append(traceback.format_exc())

    def exl2pdf_v2(self,file_name):
        try:
            if file_name[:2] == "~$":
                return

            if file_name.endswith("xls"):
                excel2 = win32.gencache.EnsureDispatch('Excel.application')
                wb = excel2.Workbooks.Open(os.path.join(self.from_path, file_name))
                xlsx_file_name = self.change_file_name_pdf(file_name, "xlsx")

                full_from_path = os.path.join(self.from_path, xlsx_file_name).replace("/", "\\")
                wb.SaveAs(full_from_path, FileFormat=51)
                wb.Close()
                excel2.Application.Quit()

                os.remove(os.path.join(self.from_path, file_name))
                file_name = xlsx_file_name



            wb = op.load_workbook(self.from_path + "/" + file_name)  # openptxl workbook생성
            ws_list = wb.sheetnames  # 해당 workbook의 시트명을 리스트로 받음
            file_name = file_name.replace(".xlsx", "")  # 파일명을 저장하기 위해 문자열에서 확장자를 제거
            result = [self.from_path + "/" + file_name, file_name, ws_list]

            excel = win32.Dispatch("Excel.Application")
            wb = excel.Workbooks.Open(result[0])  # 0번째는 파일경로
            wb.Worksheets(result[2]).Select()  # 2번쨰 요소는 시트명

            wb.ActiveSheet.ExportAsFixedFormat(0, self.to_path + "/" + result[1] + ".pdf")  # 파일명, 시트명으로 pdf 파일 저장

            wb.Close(False)  # workbook 닫기. True일 경우 그 상태를 저장한다.
            excel.Quit()  # excel 닫기
        except:
            self.error_dict["error_file"].append(file_name)
            self.error_dict["error_message"].append(traceback.format_exc())


    def exl2pdf(self):
        def excelInfo(filepath):
            excel_list = [file for file in os.listdir(filepath) if
                          file.endswith('xlsx') and file[:2] != "~$"]  # 폴더안에있는 엑셀파일 명을 리스트로 저장
            result = []  # 빈 리스트 생성

            for file in excel_list:  # 엑셀파일명 리스트를 for문을 통해 반복
                wb = op.load_workbook(filepath + "/" + file)  # openptxl workbook생성
                ws_list = wb.sheetnames  # 해당 workbook의 시트명을 리스트로 받음
                filename = file.replace(".xlsx", "")  # 파일명을 저장하기 위해 문자열에서 확장자를 제거

                for sht in ws_list:  # 시트명 리스트를 for문을 통해 반복
                    temp_tuple = (filepath + "/" + file, filename, sht)  # 파일경로, 파일명, sht를 튜플에 저장
                    result.append(temp_tuple)  # 위 튜플을 빈 리스트에 추가

            return result  # 튜플로 이루어진 리스트 리턴

        def transPDF(fileinfo, savepath):
            excel = win32.Dispatch("Excel.Application")
            i = 0  # 파일명 중복을 방지하기 위한 인덱싱 번호
            # excelinfo를 받아서 for문을 실행
            for info in fileinfo:
                wb = excel.Workbooks.Open(info[0])  # info가 튜플이므로 인덱싱으로 접근(0번째는 파일경로)
                ws = wb.Worksheets(info[2])  # 튜플의 2번쨰 요소는 시트명
                ws.Select()  # 위 설정한 시트 선택
                try:
                    wb.ActiveSheet.ExportAsFixedFormat(0, savepath + "/" + str(i) + "_" + info[1] + "_" + info[2] + ".pdf")  # 파일명, 시트명으로 pdf 파일 저장
                except:
                    pass
                i = i + 1
                wb.Close(False)  # workbook 닫기. True일 경우 그 상태를 저장한다.
                excel.Quit()  # excel 닫기

        excelinfo = excelInfo(self.from_path)
        transPDF(excelinfo, self.to_path)